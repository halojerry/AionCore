#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX 读取与数据画像（light-file-reading）。

依赖：openpyxl（读公式/格式/多 sheet），pandas（数据画像）。
铁律：openpyxl 无求值引擎——公式只存字符串。要算值用 data_only=True 读缓存，
或写后跑 LibreOffice 重算（见 references/XLSX-REF.md）。

CLI（处理真实文件）：
    python xlsx_read.py sheets   f.xlsx
    python xlsx_read.py formulas f.xlsx [--sheet Data]
    python xlsx_read.py values   f.xlsx [--sheet Data]
    python xlsx_read.py profile  f.xlsx [--sheet Data]
自检（合成 xlsx、离线、自清理）：
    python xlsx_read.py --selftest
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

_LARGE_CELLS = 500000  # 行*列超此阈值告警（解析可能慢/占内存）


def _load_wb(path, **kw):
    """惰性导入 openpyxl 并加载工作簿，缺失时给 pip 提示。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - 仅依赖缺失时触发
        raise SystemExit("缺少 openpyxl，请先安装：pip install openpyxl") from e
    return load_workbook(path, **kw)


def _warn_size(ws):
    """行列规模超阈值时告警到 stderr。"""
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    if rows * cols > _LARGE_CELLS:
        print(f"[warn] 工作表 {ws.title!r} 约 {rows} 行 x {cols} 列，"
              f"超过 {_LARGE_CELLS} 单元格，解析可能偏慢/占内存", file=sys.stderr)


def list_sheets(path):
    """返回 sheet 名列表。"""
    wb = _load_wb(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_formulas(path, sheet=None):
    """读单元格里的公式（不求值）→ [(coord, formula)]。
    sheet=None 取活动表。公式以 '=' 开头。
    用 read_only 流式读，省内存（公式字符串在 read_only 模式下仍可读）。"""
    wb = _load_wb(path, read_only=True)  # data_only 默认 False，保留公式字符串
    ws = wb[sheet] if sheet else wb.active
    _warn_size(ws)
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append((cell.coordinate, cell.value))
    wb.close()
    return out


def read_values(path, sheet=None):
    """读缓存计算值（data_only=True）→ [[...]]。
    注意：需 Excel/LibreOffice 之前算过并保存，否则公式格缓存为 None。
    用 read_only 流式读，省内存。"""
    wb = _load_wb(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    _warn_size(ws)
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return rows


def profile(path, sheet=0):
    """用 pandas 做数据画像：返回 {shape, columns, dtypes, describe}。
    sheet 可为索引或名字。"""
    pd = _imp_pandas()
    df = pd.read_excel(path, sheet_name=sheet)
    if df.shape[0] * max(df.shape[1], 1) > _LARGE_CELLS:
        print(f"[warn] sheet 约 {df.shape[0]} 行 x {df.shape[1]} 列，规模较大",
              file=sys.stderr)
    return {
        "shape": df.shape,
        "columns": list(df.columns),
        "dtypes": {c: str(t) for c, t in df.dtypes.items()},
        "describe": df.describe(include="all").to_dict(),
    }


def _imp_pandas():
    try:
        import pandas as pd
        return pd
    except ImportError as e:  # pragma: no cover - 仅依赖缺失时触发
        raise SystemExit("缺少 pandas，请先安装：pip install pandas openpyxl") from e


def _selftest():
    """合成 xlsx（含公式）跑全流程，断言后清理，离线无残留。"""
    import os
    import tempfile
    try:
        from openpyxl import Workbook
    except ImportError as e:  # pragma: no cover
        raise SystemExit("缺少 openpyxl，请先安装：pip install openpyxl") from e

    tmp = tempfile.mkdtemp(prefix="xlsxread_")
    try:
        src = os.path.join(tmp, "sample.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["n", "sq"])
        for i in range(1, 5):
            ws.append([i, i * i])
        ws["D1"] = "=SUM(A2:A5)"  # 公式格
        wb.create_sheet("Empty")
        wb.save(src)

        sheets = list_sheets(src)
        assert sheets == ["Data", "Empty"], sheets

        formulas = read_formulas(src, "Data")
        assert ("D1", "=SUM(A2:A5)") in formulas, formulas

        vals = read_values(src, "Data")
        assert vals[0][0] == "n" and vals[1][:2] == [1, 1], vals[:2]

        prof = profile(src, sheet="Data")
        assert prof["shape"] == (4, 2) and prof["columns"] == ["n", "sq"], prof
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
    print("xlsx_read self-test OK: sheets/formulas/values/profile all passed")


def _cli(argv):
    import argparse
    import json
    ap = argparse.ArgumentParser(prog="xlsx_read.py", description="XLSX 读取与数据画像")
    ap.add_argument("--selftest", action="store_true",
                    help="合成 xlsx 自检（离线、自清理），CI 用此标志")
    sub = ap.add_subparsers(dest="cmd")

    sub.add_parser("sheets", help="列出所有 sheet 名").add_argument("file")
    for name, help_ in [("formulas", "读公式（不求值）"),
                        ("values", "读缓存计算值"),
                        ("profile", "pandas 数据画像")]:
        sp = sub.add_parser(name, help=help_)
        sp.add_argument("file")
        sp.add_argument("--sheet", help="sheet 名（默认活动表/第0表）")

    args = ap.parse_args(argv)

    if args.selftest:
        _selftest()
        return
    if not args.cmd:
        ap.print_help()
        raise SystemExit(2)

    if args.cmd == "sheets":
        print(json.dumps(list_sheets(args.file), ensure_ascii=False, indent=2))
    elif args.cmd == "formulas":
        out = read_formulas(args.file, args.sheet)
        print(json.dumps(out, ensure_ascii=False, indent=2))
    elif args.cmd == "values":
        out = read_values(args.file, args.sheet)
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    elif args.cmd == "profile":
        sheet = args.sheet if args.sheet else 0
        out = profile(args.file, sheet=sheet)
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    _cli(sys.argv[1:])
